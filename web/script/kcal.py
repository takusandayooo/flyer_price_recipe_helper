import csv
import re

import openpyxl
import pandas as pd
from bs4.element import ResultSet


class Record:
    def __init__(self, number=0, name="", energy=0, protein=0, fat=0, carb=0, salt=0):
        self.number = number
        self.name = name
        self.energy = energy
        self.protein = protein
        self.fat = fat
        self.carb = carb  # carbohydrate
        self.salt = salt


def kcal(foods_list):
    return_list = []

    def sanitize(s):
        if isinstance(s, int):
            return s
        if isinstance(s, float):
            return s
        if re.match(r"\-", s) != None:
            return 0
        if re.search(r"Tr", s) != None:
            return 0
        if re.match(r"\(", s) != None:
            s = re.sub(r"\(", "", s)
            s = re.sub(r"\)", "", s)
            return float(s)
        return s

    def convert_xlsx_to_csv(uri, fn_csv):
        # データを抽出するための下準備
        wb = openpyxl.load_workbook(uri)
        print("xlsx file was opened")  # check, ok
        sn = "表全体"  # sheet name
        ws = wb[sn]  # work sheet
        datum = []
        datum.append(["食材番号", "食材名", "エネルギー", "タンパク質", "脂質", "炭水化物", "食塩相当量"])
        # データを抽出してリストに格納する。
        start, end = 13, 2490 + 1
        for j in range(start, end):
            temp = Record()
            # 食品番号、商品名
            temp.number = ws.cell(row=j, column=2).value
            temp.name = ws.cell(row=j, column=4).value
            # エネルギー [kcal]、タンパク質、脂質、炭水化物、食塩相当量
            temp.energy = sanitize(ws.cell(row=j, column=7).value)
            temp.protein = sanitize(ws.cell(row=j, column=10).value)
            temp.fat = sanitize(ws.cell(row=j, column=13).value)
            temp.carb = sanitize(ws.cell(row=j, column=21).value)
            temp.salt = sanitize(ws.cell(row=j, column=61).value)
            datum.append(
                [
                    temp.number,
                    temp.name,
                    temp.energy,
                    temp.protein,
                    temp.fat,
                    temp.carb,
                    temp.salt,
                ]
            )
        # 取得したデータをCSVファイルにする。
        with open(fn_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(datum)
        print("write csv")  # check, ok

    def main_func(foods_list, fn_csv):
        df = pd.read_csv(fn_csv, header=0, encoding="utf-8")

        # 合計栄養素値を初期化
        total_nutrients = {"エネルギー": 0, "タンパク質": 0, "脂質": 0, "炭水化物": 0, "食塩相当量": 0}

        global A
        A = [0, 0, 0, 0, 0]

        for food_name in foods_list:
            temp = df[df["食材名"].str.contains(food_name, na=False)]

            if not temp.empty:
                list_mean = Record()
                list_mean.energy = temp["エネルギー"].mean()
                list_mean.protein = temp["タンパク質"].mean()
                list_mean.fat = temp["脂質"].mean()
                list_mean.carb = temp["炭水化物"].mean()
                list_mean.salt = temp["食塩相当量"].mean()

                # Accumulate the values in list A
                A[0] += list_mean.energy
                A[1] += list_mean.protein
                A[2] += list_mean.fat
                A[3] += list_mean.carb
                A[4] += list_mean.salt

                # print("\n--- {} の統計情報 ---".format(food_name))
                # print("平均 エネルギー: {:.1f}".format(list_mean.energy))
                # print("平均 タンパク質: {:.1f}".format(list_mean.protein))
                # print("平均 脂質: {:.1f}".format(list_mean.fat))
                # print("平均 炭水化物: {:.1f}".format(list_mean.carb))
                # print("平均 食塩相当量: {:.1f}".format(list_mean.salt))
                return_list.append(
                    {
                        "name": food_name,
                        "energy": list_mean.energy,
                        "protein": list_mean.protein,
                        "fat": list_mean.fat,
                        "carb": list_mean.carb,
                        "salt": list_mean.salt,
                    }
                )
            else:
                print("\n{} のデータが見つかりませんでした".format(food_name))
                return_list.append(
                    {
                        "name": food_name,
                        "energy": 0,
                        "protein": 0,
                        "fat": 0,
                        "carb": 0,
                        "salt": 0,
                    }
                )
        # Print list A
        # print("\nList A:", A)

    # Rest of your code...

    uri = "./static/csv/20201225-mxt_kagsei-mext_01110_012.xlsx"
    fn_csv = "./static/csv/20201225-mxt_kagsei-mext_01110_012.csv"
    # uri = '20201225-mxt_kagsei-mext_01110_012.xlsx'
    # fn_csv = '20201225-mxt_kagsei-mext_01110_012.csv'
    convert_xlsx_to_csv(uri, fn_csv)
    main_func(foods_list, fn_csv)
    return return_list


if __name__ == "__main__":
    foods_list = ["米", "豚肉"]
    print(kcal(foods_list))
